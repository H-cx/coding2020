"""
实现自动更新持仓列表
持仓列表中基金总数需要手动确认或添加加仓程序
"""
from openpyxl import load_workbook
import WorthGet
def update():
    positionWorkbook = load_workbook("交易统计.xlsx")
    position = positionWorkbook.active
    codelist = ["0"]
    rowlist = 4
    columnlist = 2
    number = int(position['P4'].value)   
    while rowlist < number + 4:
        FundCode = position.cell(row = rowlist,column = columnlist).value
        name, code, netWorth, ACWorth = WorthGet.getWorth(FundCode)
        newPrice = netWorth[0]
        position.cell(row = rowlist,column = 7,value = newPrice)
        rowlist += 1
    positionWorkbook.save("交易统计.xlsx")

update()