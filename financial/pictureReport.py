import requests
import time
import execjs
import matplotlib.pyplot as plt
import Rolling
import WorthGet
from openpyxl import load_workbook

"""
用于遍历excel中基金，绘出历史净值图
已实现：
当前净值线
累计净值线
50日均线
300日均线
20日最高最低
60日最高最低
波动水平报告
突破策略信号

未实现：
50日均线与300日均线趋势数学分析
"""
def dataProcess():
    plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
    plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
    positionWorkbook = load_workbook("交易统计.xlsx")
    position = positionWorkbook.active
    codelist = ["0"]
    rowlist = 4
    columnlist = 2
    number = int(position['P4'].value)
    while rowlist < number + 4:
        FundCode = position.cell(row = rowlist,column = columnlist).value
        # 以列表形式返回净值数据
        name, code, netWorth, ACWorth = WorthGet.getWorth(FundCode)
        xmax = len(ACWorth)
        # 决策参考消息
        AC20Max = max(ACWorth[:20])
        AC20Min = min(ACWorth[:20])
        AC60Min = min(ACWorth[:60])
        AC60Max = max(ACWorth[:60])
        newWorth = ACWorth[0]
        # 20日60日突破策略信号
        if newWorth >= AC20Max and newWorth < AC60Max:
            breakMessage = "上突破20日"
        elif newWorth >= AC60Max:
            breakMessage = "上突破60日"
        elif newWorth <=AC20Min and newWorth > AC60Min:
            breakMessage = "下突破20日"
        elif newWorth <= AC60Min:
            breakMessage = "下突破60日"
        else:
            message = "暂无建议"
        # 20日与60日波动幅度策略信号
        range20 = (newWorth - AC20Min)/(AC20Max - AC20Min)    # 得到当前价格在20天内的水平
        range60 = (newWorth - AC60Min)/(AC20Max - AC60Min)    # 得到当前价格在60天内的水平
        # 倒置净值表
        netWorth = netWorth[::-1]
        ACWorth = ACWorth[::-1]
        # 数据处理
        rollingMean50 = Rolling.RollingMean(ACWorth,50)
        rollingMean300 = Rolling.RollingMean(ACWorth,300)
        # 历史净值分析
        plt.figure(figsize=(40,30))
        plt.plot(netWorth[:],color = 'k',linestyle=':',label = "最新净值")
        plt.plot(ACWorth[:],color = 'r',label = "累积净值")
        plt.plot(rollingMean50[:],color = 'b',linestyle='-.',label = "50日平均线")
        plt.plot(rollingMean300[:],color = 'm',linestyle='--',label = "300日平均线")
        # 插入20日60日参考值
        plt.text(0,1.1,"当前净值: "+str(round(newWorth,2))+"\n"+"20日最大值: "+str(round(AC20Max,2))+"\n"+"20日最小值: "+str(round(AC20Min,2))+"\n"\
            +"60日最大值: "+str(round(AC60Max,2))+"\n"+"60日最小值: "+str(round(AC60Min,2))+"\n"+"持有建议： "+message+"\n"\
                +"20日波动水平： "+ str(round(range20,2))+"\n"+"60日波动水平： "+str(round(range60,2))+"\n",fontsize = 35)
        # 做出当前净值以及20日60日的最大值最小值平均值的参考线
        plt.axhline(y=newWorth,ls = "-",c = "red",label = "当前净值")
        plt.legend()
        # 保存图片
        SavePath = "./pictures/"+name+"_"+str(code)
        plt.savefig(SavePath)
        # 输出提示信息
        print(name+str(code)+"["+str(rowlist-3)+"/"+str(number)+"]")
        rowlist += 1

# dataProcess()